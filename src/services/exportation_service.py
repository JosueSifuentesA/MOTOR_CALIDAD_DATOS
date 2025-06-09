import os
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
from pathlib import Path
from playwright.sync_api import sync_playwright
import time

def limpiar_nombre(nombre):
    
    return re.sub(r'[^A-Za-z0-9_-]', '', nombre) or "reporte"

def generar_reporte_excel(data: dict, ruta_salida: str = "./reportes") -> str:
    """
    Genera un archivo Excel presentable a partir de los datos de evaluación de calidad.
    Solo incluye los criterios seleccionados.
    """
    os.makedirs(ruta_salida, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Calidad"

    # Estilos
    color_primario = "4F95F2"
    header_fill = PatternFill(start_color=color_primario, end_color=color_primario, fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center")

    # Título
    ws.append(["Reporte de Calidad de Datos"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    ws["A1"].font = Font(size=14, bold=True, color=color_primario)
    ws["A1"].alignment = center_align

    ws.append([])  # Espacio

    # Metadatos
    data_formulario = data.get("dataFormulario", {})
    criterios_seleccionados = data_formulario.get("criterios", [])

    ws.append(["Criterios:", ", ".join(criterios_seleccionados)])
    ws.append(["Nivel de Detalle:", data_formulario.get("nivel_detalle", "")])
    ws.append(["Tipo de Reporte:", data_formulario.get("tipo_reporte", "")])
    ws.append(["Fecha de Generación:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws.append([])

    # Resumen general
    ws.append(["Resumen de Scores"])
    ws[f"A{ws.max_row}"].font = Font(size=12, bold=True, color=color_primario)
    ws.append(["Criterio", "Score"])
    for cell in ws[ws.max_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    evaluacion = data.get("resultadoEvaluacion", {})
    for criterio in criterios_seleccionados:
        clave = criterio.lower()
        if clave in evaluacion:
            score = round(evaluacion[clave].get("score", 0), 4)
            ws.append([criterio, score])

    # Score final
    ws.append([])
    ws.append(["Score Final", round(evaluacion.get("score_final", 0), 4)])
    ws[f"A{ws.max_row}"].font = Font(bold=True, color=color_primario)

    # Detalle por criterio
    for criterio in criterios_seleccionados:
        clave = criterio.lower()
        if clave not in evaluacion:
            continue

        ws.append([])
        ws.append([f"Detalle: {criterio}"])
        ws[f"A{ws.max_row}"].font = Font(size=12, bold=True, color=color_primario)

        ws.append(["Campo", "Score"])
        for cell in ws[ws.max_row]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align

        detalle = evaluacion[clave].get("detalle", {})
        for campo, score in detalle.items():
            ws.append([campo, round(score, 4)])

    # Ajustar anchos
    for i, col in enumerate(ws.columns, start=1):
        max_length = 0
        col_letter = get_column_letter(i)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    # Guardar
    nombre_base = limpiar_nombre(data_formulario.get("nombre_archivo", "reporte_calidad"))
    ruta_completa = os.path.join(ruta_salida, f"{nombre_base}.xlsx")
    wb.save(ruta_completa)

    return ruta_completa

def export_html_to_pdf_via_http(temp_filename, pdf_output_path):
    from playwright.sync_api import sync_playwright

    with sync_playwright() as p:
        browser = p.chromium.launch()
        page = browser.new_page()

        # Cargar el HTML alojado en /static/temp/
        page.goto(f"http://localhost:5000/static/temp/{temp_filename}", wait_until="networkidle")

        # Esperar que se cargue todo el contenido
        page.wait_for_timeout(2000)

        # Obtener la altura total del contenido renderizado (en px)
        height_px = page.evaluate("document.body.scrollHeight")
        # Convertir a pulgadas
        height_in = height_px / 96  # 96 px = 1 in

        # Asegurarse que el alto mínimo sea al menos A4 (11.7in)
        height_in = max(height_in, 11.7)

        # Exportar a PDF con altura dinámica
        page.pdf(
            path=str(pdf_output_path),
            format="A4",
            print_background=True,
            prefer_css_page_size=True
        )

        browser.close()





